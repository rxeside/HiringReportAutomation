import os
import asyncio
import openpyxl
from huntflow_api_client import HuntflowAPI
from huntflow_api_client.tokens.token import ApiToken
import httpx
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, timezone

HUNTFLOW_API_TOKEN = "a5ac927e15662b47b55c0f857c499cf805b543e29532a586d8ef3e384e14e605"

OUTPUT_FILE_NAME = "voronka_kandidatov_final.xlsx"

PRIORITY_VACANCIES = {
    "РОП  NA AM",
    "Директор по персоналу 2025",
    "Руководитель проектов в маркетинг",
    "Менеджер по прогреву воронки",
    "Тимлид проджектов в PD",
}

FUNNEL_STAGES_ORDER = [
    "коннект",
    "интервью с HR",
    "интервью с заказчиком",
    "финальное интервью",
    "выставлен оффер",
    "вышел на работу",
]

HUNTFLOW_STATUSES_TO_COLUMNS = {
    "Коннект": "коннект",
    "Интервью с HR": "интервью с HR",
    "Интервью с заказчиком": "интервью с заказчиком",
    "Финальное интервью": "финальное интервью",
    "Предложение о работе": "выставлен оффер",
    "Вышел на работу": "вышел на работу",
}


async def get_total_applicants_on_stage(api_client, account_id, vacancy_id, status_id):
    """(Для чисел вне скобок) Получает ОБЩЕЕ число кандидатов на этапе."""
    try:
        search_params = {"vacancy": [vacancy_id], "status": [status_id], "only_current_status": "false", "count": 1}
        response = await api_client.request("GET", f"/accounts/{account_id}/applicants/search", params=search_params)
        data = response.json()
        return data.get("total_items", 0)
    except Exception as e:
        print(f"      - Ошибка при получении общего числа для status_id {status_id}: {e}")
        return 0


async def get_factual_weekly_funnel_counts(api_client, account_id, vacancy_id, status_id_to_name_map):
    """
    Анализирует логи, чтобы найти ФАКТИЧЕСКИ пройденные этапы ЗА ПОСЛЕДНЮЮ НЕДЕЛЮ.
    Учитывает "перепрыгивания" через этапы.
    """
    weekly_factual_counts = {stage: 0 for stage in FUNNEL_STAGES_ORDER}
    all_applicants = []
    try:
        page = 1
        while True:
            search_params = {"vacancy": [vacancy_id], "only_current_status": "false", "count": 100, "page": page}
            response = await api_client.request("GET", f"/accounts/{account_id}/applicants/search",
                                                params=search_params)
            data = response.json()
            items = data.get('items', [])
            if not items: break
            all_applicants.extend(items)
            page += 1
    except Exception as e:
        print(f"      - Ошибка при получении списка кандидатов для анализа логов: {e}")
        return weekly_factual_counts

    one_week_ago = datetime.now(timezone.utc) - timedelta(days=7)

    for applicant in all_applicants:
        applicant_id = applicant.get("id")
        counted_stages_for_applicant = set()
        try:
            log_params = {"vacancy": vacancy_id}
            log_response = await api_client.request("GET", f"/accounts/{account_id}/applicants/{applicant_id}/logs",
                                                    params=log_params)
            logs = list(reversed(log_response.json().get("items", [])))

            for i, log in enumerate(logs):
                if log.get("type") == "STATUS":
                    status_name = status_id_to_name_map.get(log.get("status"))
                    column_name = HUNTFLOW_STATUSES_TO_COLUMNS.get(status_name)
                    log_date_str = log.get("created")

                    if not (column_name and log_date_str): continue

                    log_date = datetime.fromisoformat(log_date_str)
                    if log_date < one_week_ago: continue

                    if (i + 1) < len(logs) and logs[i + 1].get("type") == "COMMENT":
                        stage_index = FUNNEL_STAGES_ORDER.index(column_name)
                        stages_to_count = FUNNEL_STAGES_ORDER[0:stage_index + 1]

                        for stage in stages_to_count:
                            if stage not in counted_stages_for_applicant:
                                weekly_factual_counts[stage] += 1
                                counted_stages_for_applicant.add(stage)
        except Exception:
            continue
    return weekly_factual_counts


async def get_huntflow_data(api_client):
    try:
        accounts_response = await api_client.request("GET", "/accounts")
        accounts_data = accounts_response.json()
        if not accounts_data.get("items"): print("Ошибка: Не найдено ни одного аккаунта."); return None
        account_id = accounts_data["items"][0]["id"]
        print(f"Успешно подключились к аккаунту: {accounts_data['items'][0]['name']} (ID: {account_id})")

        statuses_response = await api_client.request("GET", f"/accounts/{account_id}/vacancies/statuses")
        statuses_data = statuses_response.json()
        status_name_to_id_map = {status["name"]: status["id"] for status in statuses_data.get("items", [])}
        status_id_to_name_map = {v: k for k, v in status_name_to_id_map.items()}

        vacancies_response = await api_client.request("GET", f"/accounts/{account_id}/vacancies",
                                                      params={"opened": "true"})
        vacancies_data = vacancies_response.json()
        all_vacancies_data = []
        print(f"\nНайдено {len(vacancies_data.get('items', []))} активных вакансий. Начинаю сбор данных...")

        for vacancy in vacancies_data.get("items", []):
            vacancy_id = vacancy["id"]
            vacancy_position = vacancy["position"]

            # заглушка для поиска только приоритетных
            # if vacancy_position not in PRIORITY_VACANCIES:
            #     continue

            print(f"  - Обрабатываю вакансию: «{vacancy_position}»")

            is_priority = vacancy_position in PRIORITY_VACANCIES
            funnel_row = {"название вакансии": vacancy_position, "комментарий": "", "is_priority": is_priority}
            for column_name in FUNNEL_STAGES_ORDER:
                funnel_row[column_name] = {"total": 0, "current": 0}

            weekly_counts = await get_factual_weekly_funnel_counts(api_client, account_id, vacancy_id,
                                                                   status_id_to_name_map)
            for stage_name, count in weekly_counts.items():
                funnel_row[stage_name]["current"] = count

            for status_hf_name, column_name in HUNTFLOW_STATUSES_TO_COLUMNS.items():
                status_id = status_name_to_id_map.get(status_hf_name)
                if status_id:
                    total_count = await get_total_applicants_on_stage(api_client, account_id, vacancy_id, status_id)
                    if column_name in funnel_row:
                        funnel_row[column_name]["total"] = total_count
            all_vacancies_data.append(funnel_row)

        all_vacancies_data.sort(key=lambda x: not x.get('is_priority', False))
        return all_vacancies_data

    except httpx.HTTPStatusError as e:
        if e.response.status_code == 401:
            print("Ошибка 401: Unauthorized. Токен недействителен или истек.")
        else:
            print(f"Произошла HTTP ошибка: {e.response.status_code} - {e.response.text}")
        return None
    except Exception as e:
        import traceback
        print(f"Произошла непредвиденная ошибка: {e}")
        traceback.print_exc()
        return None


def create_xlsx_report(data):
    if not data: print("Нет данных для создания отчета."); return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Воронка кандидатов"
    priority_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    headers = ["название вакансии"] + FUNNEL_STAGES_ORDER + ["комментарий"]
    sheet.append(headers)

    for row_data in data:
        row_to_append = []
        for header in headers:
            cell_data = row_data.get(header)
            if isinstance(cell_data, dict):
                total = cell_data.get('total', 0)
                current = cell_data.get('current', 0)
                formatted_value = f"{total} ({current})"
                row_to_append.append(formatted_value)
            else:
                row_to_append.append(cell_data if cell_data is not None else "")
        sheet.append(row_to_append)
        if row_data.get('is_priority', False):
            for cell in sheet[sheet.max_row]:
                cell.fill = priority_fill

    try:
        workbook.save(OUTPUT_FILE_NAME)
        print(f"\nОтчет успешно сохранен в файл: {os.path.abspath(OUTPUT_FILE_NAME)}")
    except Exception as e:
        print(f"Не удалось сохранить файл. Ошибка: {e}")


async def main():
    api_client = HuntflowAPI(base_url="https://api.huntflow.ru", token=ApiToken(access_token=HUNTFLOW_API_TOKEN))
    funnel_data = await get_huntflow_data(api_client)
    if funnel_data:
        create_xlsx_report(funnel_data)


if __name__ == "__main__":
    asyncio.run(main())