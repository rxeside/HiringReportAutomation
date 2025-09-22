import asyncio
import openpyxl
from huntflow_api_client import HuntflowAPI
from huntflow_api_client.tokens.token import ApiToken
import httpx
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, timezone
import logging
from io import BytesIO
from typing import Dict, Any, List, Optional
import traceback
from .token_manager import token_proxy

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

PRIORITY_VACANCIES = {
    "РОП  NA AM", "Директор по персоналу 2025", "Руководитель проектов в маркетинг",
    "Менеджер по прогреву воронки", "Тимлид проджектов в PD",
}
FUNNEL_STAGES_ORDER = [
    "коннект", "интервью с HR", "интервью с заказчиком", "финальное интервью",
    "выставлен оффер", "вышел на работу",
]
HUNTFLOW_STATUSES_TO_COLUMNS = {
    "Коннект": "коннект", "Интервью с HR": "интервью с HR",
    "Интервью с заказчиком": "интервью с заказчиком", "Финальное интервью": "финальное интервью",
    "Предложение о работе": "выставлен оффер", "Вышел на работу": "вышел на работу",
}


async def _fetch_all_paginated_items(api_client: HuntflowAPI, url: str, params: Dict = None) -> List[Dict]:
    all_items = []
    current_page = 1
    total_pages = 1
    base_params = params.copy() if params else {}

    while current_page <= total_pages:
        base_params["page"] = current_page
        base_params["count"] = 100
        response = await api_client.request("GET", url, params=base_params)
        data = response.json()
        items = data.get("items", [])
        if not items:
            break
        all_items.extend(items)
        if current_page == 1:
            total_pages = data.get("total_pages", 1)
        current_page += 1
    return all_items


async def _process_applicant_logs(api_client: HuntflowAPI, account_id: int, applicant: Dict, vacancy_id: int, status_id_to_name_map: Dict) -> Dict:
    weekly_counts = {stage: 0 for stage in FUNNEL_STAGES_ORDER}
    applicant_id = applicant.get("id")
    counted_stages = set()
    one_week_ago = datetime.now(timezone.utc) - timedelta(days=7)

    try:
        logs_url = f"/accounts/{account_id}/applicants/{applicant_id}/logs"
        log_response = await api_client.request("GET", logs_url, params={"vacancy": vacancy_id})
        logs = list(reversed(log_response.json().get("items", [])))

        for i, log in enumerate(logs):
            if log.get("type") != "STATUS":
                continue

            log_date = datetime.fromisoformat(log.get("created", ""))
            if log_date < one_week_ago:
                continue

            status_name = status_id_to_name_map.get(log.get("status"))
            column_name = HUNTFLOW_STATUSES_TO_COLUMNS.get(status_name)
            if not column_name or (i + 1) >= len(logs) or logs[i + 1].get("type") != "COMMENT":
                continue

            stage_index = FUNNEL_STAGES_ORDER.index(column_name)
            for stage in FUNNEL_STAGES_ORDER[0:stage_index + 1]:
                if stage not in counted_stages:
                    weekly_counts[stage] += 1
                    counted_stages.add(stage)
    except Exception as e:
        logging.warning(f"Ошибка при обработке логов кандидата {applicant_id} для вакансии {vacancy_id}: {e}")

    return weekly_counts


async def _build_funnel_row(api_client: HuntflowAPI, account_id: int, vacancy: Dict, status_maps: Dict) -> Dict:
    vacancy_position = vacancy.get("position", "Без названия")
    vacancy_id = vacancy["id"]
    member_ids = await get_vacancy_coworkers(api_client, account_id, vacancy_id)

    funnel_row = {"название вакансии": vacancy_position, "is_priority": vacancy_position in PRIORITY_VACANCIES,
                  "members": member_ids}
    for column_name in FUNNEL_STAGES_ORDER:
        funnel_row[column_name] = {"total": 0, "current": 0}

    weekly_counts = await get_factual_weekly_funnel_counts(api_client, account_id, vacancy_id,
                                                           status_maps['id_to_name'])
    for stage_name, count in weekly_counts.items():
        funnel_row[stage_name]["current"] = count

    for status_hf, column in HUNTFLOW_STATUSES_TO_COLUMNS.items():
        status_id = status_maps['name_to_id'].get(status_hf)
        if status_id:
            total_count = await get_total_applicants_on_stage(api_client, account_id, vacancy_id, status_id)
            funnel_row[column]["total"] = total_count

    return funnel_row


async def get_vacancy_coworkers(api_client: HuntflowAPI, account_id: int, vacancy_id: int) -> List[int]:
    try:
        params = {
            "vacancy_id": [vacancy_id],
            "type": ["owner", "manager"]
        }
        response = await api_client.request("GET", f"/accounts/{account_id}/coworkers", params=params)
        data = response.json()
        return [item['id'] for item in data.get("items", [])]
    except Exception as e:
        logging.error(f"Не удалось получить рекрутеров для вакансии {vacancy_id}: {e}")
        return []


async def get_coworkers(api_client: HuntflowAPI, account_id: int) -> Dict[int, str]:
    try:
        url = f"/accounts/{account_id}/coworkers"
        all_coworkers_items = await _fetch_all_paginated_items(api_client, url)
        coworkers_map = {item["id"]: item["name"] for item in all_coworkers_items}
        logging.info(f"Успешно загружено {len(coworkers_map)} рекрутеров (общий список).")
        return coworkers_map
    except Exception as e:
        logging.error(f"Произошла ошибка при получении общего списка рекрутеров: {e}")
        return {}


async def get_total_applicants_on_stage(api_client: HuntflowAPI, account_id: int, vacancy_id: int, status_id: int) -> int:
    try:
        url = f"/accounts/{account_id}/applicants/search"
        params = {"vacancy": [vacancy_id], "status": [status_id], "only_current_status": "false", "count": 1}
        response = await api_client.request("GET", url, params=params)
        return response.json().get("total_items", 0)
    except Exception as e:
        logging.error(f"Ошибка при получении общего числа для status_id {status_id}: {e}")
        return 0


async def get_factual_weekly_funnel_counts(api_client: HuntflowAPI, account_id: int, vacancy_id: int, status_id_to_name_map: Dict) -> Dict:
    weekly_factual_counts = {stage: 0 for stage in FUNNEL_STAGES_ORDER}
    applicants_url = f"/accounts/{account_id}/applicants/search"
    all_applicants = await _fetch_all_paginated_items(api_client, applicants_url, params={"vacancy": [vacancy_id]})

    for applicant in all_applicants:
        applicant_counts = await _process_applicant_logs(api_client, account_id, applicant, vacancy_id, status_id_to_name_map)
        for stage, count in applicant_counts.items():
            weekly_factual_counts[stage] += count

    return weekly_factual_counts


async def generate_recruitment_funnel_report() -> Optional[Dict[str, Any]]:
    if not token_proxy._access_token:
        logging.error("Токен Huntflow не предоставлен.")
        return None

    api_client = HuntflowAPI(
        base_url="https://api.huntflow.ru",
        token_proxy=token_proxy,
        auto_refresh_tokens=True
    )

    try:
        accounts_response = await api_client.request("GET", "/accounts")
        account_id = accounts_response.json()["items"][0]["id"]
        logging.info(f"Успешно подключились к аккаунту ID: {account_id}")

        coworkers_map = await get_coworkers(api_client, account_id)

        statuses_response = await api_client.request("GET", f"/accounts/{account_id}/vacancies/statuses")
        statuses_items = statuses_response.json().get("items", [])
        status_maps = {
            'name_to_id': {status["name"]: status["id"] for status in statuses_items},
            'id_to_name': {status["id"]: status["name"] for status in statuses_items}
        }

        vacancies_url = f"/accounts/{account_id}/vacancies"
        all_vacancies = await _fetch_all_paginated_items(api_client, vacancies_url, params={"opened": "true"})
        logging.info(f"Найдено {len(all_vacancies)} активных вакансий. Начинаю сбор данных...")

        tasks = [_build_funnel_row(api_client, account_id, v, status_maps) for v in all_vacancies if
                 v.get("position") in PRIORITY_VACANCIES]
        all_vacancies_data = await asyncio.gather(*tasks)

        all_vacancies_data.sort(key=lambda x: not x.get('is_priority', False))
        return {"vacancies": all_vacancies_data, "coworkers": coworkers_map}
    except Exception as e:
        logging.error(f"КРИТИЧЕСКАЯ ОШИБКА в generate_recruitment_funnel_report: {e}")
        logging.error(traceback.format_exc())
        return None


def create_xlsx_report(data: List[Dict]) -> Optional[BytesIO]:
    if not data:
        logging.warning("Нет данных для создания отчета.")
        return None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Воронка кандидатов"
    priority_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    headers = ["название вакансии"] + FUNNEL_STAGES_ORDER + ["комментарий"]
    sheet.append(headers)

    for row_data in data:
        row_to_append = []
        for header in headers:
            cell_data = row_data.get(header)
            if isinstance(cell_data, dict):
                row_to_append.append(f"{cell_data.get('total', 0)} ({cell_data.get('current', 0)})")
            else:
                row_to_append.append(cell_data if cell_data is not None else "")
        sheet.append(row_to_append)
        if row_data.get('is_priority', False):
            for cell in sheet[sheet.max_row]:
                cell.fill = priority_fill

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook